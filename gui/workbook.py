#    General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#    model for 1st, 2nd and 3rd generation solar cells.
#    Copyright (C) 2012 Roderick C. I. MacKenzie <r.c.i.mackenzie@googlemail.com>
#
#	www.gpvdm.com
#	Room B86 Coates, University Park, Nottingham, NG7 2RD, UK
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License v2.0, as published by
#    the Free Software Foundation.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.

from dat_file_class import dat_file

work_book_enabled=False
try:
	from openpyxl import Workbook
	from openpyxl.chart import Reference, Series, ScatterChart
	from openpyxl.compat import range
	work_book_enabled=True
except:
	print("python3-openpyxl not found")

import glob
import os
from plot_io import plot_load_info
from dat_file import dat_file_read
from dat_file import dat_file

from help import help_window

def title_truncate(title):
	ret=title

	if len(ret)>30:
		ret=ret[:30]

	return ret

def gen_workbook(input_file_or_dir,output_file):
	if work_book_enabled==False:
		return

	wb = Workbook()
	if os.path.isfile(input_file_or_dir):
		files=[input_file_or_dir]
	if os.path.isdir(input_file_or_dir):
		files=glob.glob(os.path.join(input_file_or_dir,"*.dat"))
	else:
		return
	
	for my_file in files:
		#print("about to save1",my_file)
		print(my_file)
		data=dat_file()
		if dat_file_read(data,my_file,guess=False)==True:
			x=[]
			y=[]
			z=[]
			if dat_file_read(data,my_file)==True:
				#print("read",my_file)
				ws = wb.create_sheet(title=title_truncate(os.path.basename(my_file)))
				ws.cell(column=1, row=1, value=data.title)
				ws.cell(column=1, row=2, value=data.x_label+" ("+data.x_units+") ")
				ws.cell(column=2, row=2, value=data.data_label+" ("+data.data_units+") ")
		
				for i in range(0,data.y_len):
					ws.cell(column=1, row=i+3, value=data.y_scale[i])
					ws.cell(column=2, row=i+3, value=data.data[0][0][i])

				c1 = ScatterChart()
				c1.title = data.title
				c1.style = 13
				c1.height=20
				c1.width=20
				c1.y_axis.title = data.data_label+" ("+data.data_units+") "
				c1.x_axis.title = data.x_label+" ("+data.x_units+") "

				xdata = Reference(ws, min_col=1, min_row=3, max_row=3+data.y_len)
				ydata = Reference(ws, min_col=2, min_row=3, max_row=3+data.y_len)

				series = Series(ydata,xdata,  title_from_data=True)
				c1.series.append(series)
				ws.add_chart(c1, "G4")
		print(my_file)
	#print("about to save1")
	try:
		wb.save(filename = output_file)
	except:
		return False
	print("exit")
	return True
	#print("about to save0")
