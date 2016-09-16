#    General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#    model for 1st, 2nd and 3rd generation solar cells.
#    Copyright (C) 2012 Roderick C. I. MacKenzie <r.c.i.mackenzie@googlemail.com>
#
#	www.gpvdm.com
#	Room B86 Coates, University Park, Nottingham, NG7 2RD, UK
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License v2.0, as published by
#    the Free Software Foundation.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.


import os
from tab import tab_class
from window_list import windows
from cal_path import get_image_file_path

#qt
from PyQt5.QtCore import QSize, Qt 
from PyQt5.QtWidgets import QWidget,QVBoxLayout,QToolBar,QSizePolicy,QAction,QTableWidget,QAbstractItemView
from PyQt5.QtGui import QPainter,QIcon

#python modules
import webbrowser

from help import help_window

from epitaxy import epitaxy_get_layers
from epitaxy import epitaxy_get_width
from mesh import mesh_get_xlen
from mesh import mesh_get_zlen
from epitaxy import epitaxy_get_mat_file
from openpyxl import Workbook

from gui_util import tab_add

from openpyxl import load_workbook
from cal_path import get_materials_path


articles = []
mesh_articles = []

class cost(QWidget):

	def callback_close(self, widget, data=None):
		self.hide()
		return True

	def callback_help(self):
		webbrowser.open('http://www.gpvdm.com/man/index.html')

	def __init__(self):
		QWidget.__init__(self)
		self.setFixedSize(900, 600)
		self.setWindowIcon(QIcon(os.path.join(get_image_file_path(),"jv.png")))

		self.setWindowTitle(_("Cost and energy payback calculator (BETA - missing realistic data at the moment!!!)")) 
		

		self.main_vbox = QVBoxLayout()

		toolbar=QToolBar()
		toolbar.setIconSize(QSize(48, 48))

		self.play = QAction(QIcon(os.path.join(get_image_file_path(),"play.png")), "Re-calcualte", self)
		self.play.triggered.connect(self.update)
		toolbar.addAction(self.play)
		
		spacer = QWidget()
		spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
		toolbar.addWidget(spacer)


		self.help = QAction(QIcon(os.path.join(get_image_file_path(),"help.png")), "Help", self)
		self.help.triggered.connect(self.callback_help)
		toolbar.addAction(self.help)

		self.main_vbox.addWidget(toolbar)


		self.tab= QTableWidget()

		self.main_vbox.addWidget(self.tab)



		self.setLayout(self.main_vbox)
		self.win_list=windows()
		self.win_list.load()
		self.win_list.set_window(self,"costs_window")

		self.update()

	def update(self):
		self.tab.clear()
		self.tab.setColumnCount(5)
		self.tab.setRowCount(0)
		self.tab.setSelectionBehavior(QAbstractItemView.SelectRows)
		self.tab.setHorizontalHeaderLabels([_("material"), _("Volume (m^-3)"), _("Mass (kg)"), _("Cost ($)"), _("Energy (J)")])
		self.tab.setColumnWidth(1, 200)

		energy_tot=0.0
		cost_tot=0.0
		for i in range(0,epitaxy_get_layers()):
			
			volume=epitaxy_get_width(i)*1.0*1.0
			name=epitaxy_get_mat_file(i)
			xls_file_name=os.path.join(get_materials_path(),epitaxy_get_mat_file(i),"cost.xlsx")
			wb = load_workbook(xls_file_name)
			ws= wb.get_sheet_by_name("results")

			density = float(ws['B2'].value)
			mass=density*volume

			cost_per_kg = float(ws['B3'].value)
			cost=mass*cost_per_kg

			energy_per_kg = float(ws['B4'].value)
			energy=energy_per_kg*mass

			tab_add(self.tab,[name,str(volume),str(mass),str(cost),str(energy)])

			energy_tot=energy_tot+energy
			cost_tot=cost_tot+cost
		
		
		payback_time=1.0
		tab_add(self.tab,["sum","","",str(cost_tot),str(energy_tot)])
		tab_add(self.tab,["","","pay back time=",str(payback_time),"years"])
		
	def closeEvent(self, event):
		self.win_list.update(self,"costs_window")
		self.hide()
		event.accept()



