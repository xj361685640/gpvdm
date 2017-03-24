# -*- coding: utf-8 -*-
#    General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#    model for 1st, 2nd and 3rd generation solar cells.
#    Copyright (C) 2017 Edward Grant  eayeg3 at nottingham.ac.uk
#    Copyright (C) 2017 Roderick C. I. MacKenzie r.c.i.mackenzie at googlemail.com
#
#	https://www.gpvdm.com
#	Room B86 Coates, University Park, Nottingham, NG7 2RD, UK
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License v2.0, as published by
#    the Free Software Foundation.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.

import sys

import numpy as np
import datetime
import os
import openpyxl

from spectrum_pref import *
from spectrum_planet import *
from spectrum_solargeometry import *

class earthModel(object):
    def __init__(self, Latitude, Longitude, W, p, Date, Time, AOD, Timezone):
        self.Latitude = Latitude
        self.Longitude = Longitude
        self.W = W
        self.p = p
        self.Date = Date
        self.Time = Time
        self.AOD = AOD
        self.Timezone = Timezone


def earth_calc(Latitude, Longitude, W, p, Date, Time, AOD, timezone):
    # BLACK BODY MODEL - equation easily found online

    # define constants and variables
    lam_bb = np.linspace(300e-9, 4000e-9, 1000)  # setting array of wavelengths
    h = 6.626e-34  # planck's constant
    c = 3e8  # speed of light
    T = 5800  # temp of sun
    k = 1.38066e-23  # boltzmann constant
    r_sun = 695700000  # sun's radius
    AU = 149597870700  # 1 astronomical unit
    r_earth_orbit = 1 * AU  # earth's orbital radius

    # spectral radiance
    sol = ((2 * h * c * c) / lam_bb ** 5) * 1 / (np.exp((h * c) / (lam_bb * k * T)) - 1)

    # multiply by the square of the ratio of the solar radius of earth's orbital radius
    sol = sol * (r_sun / r_earth_orbit) ** 2

    # lambert's cosine law
    sol = sol * np.pi

    sol = 1e-9 * sol  # units are now Wm-2nm-1

    # EARTH MODEL - model used is called SPCTRAL2 by R. Bird and C. Riordan. Scientific paper outlining equations can
    # be seen in .../Papers and Sources/spctral2.pdf

    # Variables and constants

    # pulling values from Date and Time user inputted variables
    day = int(Date[0:2])
    month = int(Date[3:5])
    year = int(Date[6:10])
    hour = int(Time[0:2])
    minute = int(Time[3:5])

    d = datetime.datetime(year, month, day, hour, minute).timetuple().tm_yday  # day of the year
    p_o = 1013.25  # constant
    alpha = 1.14  # from angstrom expression, used in aerosol scattering and absorbing

    # solar zenith angle in terms of location and time
    # altitude = get_altitude(Latitude, Longitude, datetime.datetime(year, month, day, hour, minute, 0, 0))
    # Z_deg = 90 - altitude
    Z_deg = get_zenith(Latitude, Longitude, d, hour, minute, timezone)
    Z_rad = Z_deg * math.pi / 180  # in rads

    # calling up location of excel file
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','solar_data.xlsx')

    # reading columns A:F and converting them to values
    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A2:F123']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals = np.array(vals_array)

    # pulling wavelength data
    lam = vals[0:122, 0]

    # extra-terrestrial radiation at these wavelengths
    ext_ter_spec = vals[0:122, 1]

    # Earth-Sun Correction factor
    psi = 2 * math.pi * (d - 1) / 365
    D = 1.00011 + 0.034221 * math.cos(psi) + 0.00128 * math.sin(psi) + 0.00719 * math.cos(
        2 * psi) + 0.000077 * math.sin(2 * psi)

    # T_rlam: Rayleigh Scattering
    M = (math.cos(Z_rad) + 0.15 * (93.885 - Z_deg) ** (-1.253)) ** (-1)  # relative air mass
    M_ = M * (p / p_o)  # pressure corrected air mass
    T_rlam = np.exp((-M_) / ((lam ** 4) * (115.6406 - 1.3366 / (lam ** 2))))

    # T_alam: Aerosol Scattering and absorption
    T_alam = np.exp(-AOD * M * (lam / 0.5) ** (-alpha))

    # T_wlam: Water Vapour Absorption
    a_wlam = vals[0:122, 2]  # water vapour absorption coefficients
    T_wlam = np.exp((-0.2385 * a_wlam * W * M) / (1 + 20.07 * a_wlam * W * M) ** 0.45)

    # T_olam: Ozone and uniformly mixed gas absorption
    a_o3lam = vals[0:122, 3]  # ozone absorption coefficients
    M_o = (1 + 22 / 6370) / (math.cos(Z_rad) ** 2 + 2 * 22 / 6370) ** 0.5  # effective ozone mass
    T_olam = np.exp(-a_o3lam * 0.395 * M_o)

    a_ulam = vals[0:122, 4]  # uniformly mixed gas absorption coefficients
    T_ulam = np.exp((-1.41 * a_ulam * M_) / (1 + 118.93 * a_ulam * M_) ** 0.45)

    # Diffuse irradiance on a horizontal surface
    # Rayleigh Scattering Component
    C_s = vals[0:122, 5]
    OMEGL = 0.945 * np.exp(-0.095 * (np.log(lam / 0.4)) ** 2)
    DELA = AOD * ((lam / 0.5) ** -1.14)
    T_aalam = np.exp(-(1 - OMEGL) * DELA * M)
    I_rlam = ext_ter_spec * D * math.cos(Z_rad) * T_olam * T_ulam * T_wlam * T_aalam * (1 - T_rlam ** 0.95) * 0.5 * C_s

    # Aerosol Scattering Component
    ALG = math.log(1 - 0.65)
    BFS = ALG * (0.0783 + ALG * (-0.3824 - ALG * 0.5874))
    AFS = ALG * (1.459 + ALG * (0.1595 + ALG * 0.4129))
    F_s = 1 - 0.5 * math.exp((AFS + BFS * math.cos(Z_rad)) * math.cos(Z_rad))
    T_aslam = np.exp(-OMEGL * DELA * M)
    I_alam = ext_ter_spec * D * math.cos(Z_rad) * T_olam * T_ulam * T_wlam * T_aalam * (T_rlam ** 1.5) * (
    1 - T_aslam) * F_s * C_s


    # Irradiance
    I_direct = ext_ter_spec * D * T_rlam * T_alam * T_wlam * T_olam * T_ulam

    I_diffuse = I_rlam + I_alam  # +I_glam

    I_total = I_direct + I_diffuse

    # Solar Constant Calc (using trapezium rule to integrate I_total over wavelengths)
    solar_const = 0

    for i in range(1,len(I_total)):
        solar_const = solar_const + 0.5*(lam[i]-lam[i-1])*(I_total[i]+I_total[i-1])

    #Luminosity
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','luminosityFunction.xlsx')

    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A1:B122']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals1 = np.array(vals_array)

    lumin = vals1[0:122,1]

    #luminosity

    luminosity = 0

    for i in range(0, len(lam)):
        luminosity = luminosity + 683*10**(-3)*lumin[i]*I_total[i]

    print(luminosity)

    return lam, I_direct, I_diffuse, I_total, lam_bb, sol, ext_ter_spec, solar_const

def black_body(a, ecc, orbitalpoint):
    # BLACK BODY MODEL
    # define constants and variables
    lam_bb = np.linspace(300e-9, 4000e-9, 371)  # setting array of wavelengths
    h = 6.626e-34  # planck's constant
    c = 3e8  # speed of light
    T = 5800  # temp of sun
    k = 1.38066e-23  # boltzmann constant
    r_sun = 695700000  # sun's radius
    AU = 149597870700  # 1 astronomical unit
    a = a*AU

    # Equation for r is the equation of an elliptical orbit, knowing the eccentricity, and the semi-major axis (a)
    r = a*(1-ecc**2)/(1+ecc*math.cos(orbitalpoint*2*math.pi))

    # spectral radiance
    sol = ((2 * h * c * c) / lam_bb ** 5) * 1 / (np.exp((h * c) / (lam_bb * k * T)) - 1)

    # multiply by the square of the ratio of the solar radius of earth's orbital radius
    sol = sol * (r_sun / r) ** 2

    # lambert's cosine law
    sol = sol * np.pi

    sol = 1e-9 * sol  # units are now Wm-2nm-1

    #EXTRA TERRESTRIAL
    # calling up location of excel file
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','solar_data.xlsx')

    # reading columns A:F and converting them to values
    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A2:F123']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals = np.array(vals_array)

    # pulling wavelength data
    lam = vals[0:122, 0]

    # extra-terrestrial radiation at these wavelengths
    ext_ter_spec = vals[0:122, 1]

    #Correcting for distance from sun
    ext_ter_spec = ext_ter_spec*(AU**2/r**2)

    maximum = np.amax(ext_ter_spec)

    # Solar Constant Calc (using trapezium rule to integrate I_total over wavelengths
    solar_const = 0

    for i in range(1,len(lam)):
        solar_const = solar_const + 0.5 * (lam[i] - lam[i - 1]) * (ext_ter_spec[i] + ext_ter_spec[i - 1])

    return sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam

def orbital_plot_calc(a, ecc, orbitalpoint):
    AU = 149597870700  # 1 astronomical unit
    a = a*AU
    theta = np.linspace(0, 2 * math.pi, num=1000)

    r_plot = a * (1 - ecc ** 2) / (1 + ecc * np.cos(theta))
    r2 = a*(1-ecc**2)/(1+ecc*math.cos(orbitalpoint*2*math.pi))

    return r_plot, theta, r2

def mars_calc(orbitalpoint,AOD):
    ecc = 0.0934
    a = 1.523679

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)


    # T_alam: Aerosol Scattering and absorption
    Z_rad = 0
    Z_deg = 0
    M = (math.cos(Z_rad) + 0.15 * (93.885 - Z_deg) ** (-1.253)) ** (-1)  # relative air mass
    M=1
    T_alam = np.exp(-AOD * M * (lam / 0.5) ** (-1.14))

    #CO2 absorption
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','co2 data.xlsx')

    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A1:B122']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals = np.array(vals_array)

    T_CO2 = vals[0:122,1]


    total = ext_ter_spec*T_CO2*T_alam

    solar_const = 0

    for i in range(1,len(lam)):
        solar_const = solar_const + 0.5 * (lam[i] - lam[i - 1]) * (total[i] + total[i - 1])

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam, total

def mercury_calc(orbitalpoint):
    ecc = 0.20563
    a = 0.387098

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam

def venus_calc(orbitalpoint):
    ecc = 0.006772
    a = 0.723332

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    #CO2 absorption
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','co2 data.xlsx')

    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A1:B122']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals = np.array(vals_array)

    T_CO2 = vals[0:122,1]


    #total
    total = 0.25*ext_ter_spec*T_CO2

    solar_const = 0

    for i in range(1, len(lam)):
        solar_const = solar_const + 0.5 * (lam[i] - lam[i - 1]) * (total[i] + total[i - 1])

    #Luminosity
    file_location = os.path.join(os.getcwd(), 'materials', 'gasses','luminosityFunction.xlsx')

    wb = openpyxl.load_workbook(filename=file_location)
    sheet_ranges = wb['Sheet1']
    vals = sheet_ranges['A1:B122']
    vals_array = []
    for row in vals:
        vals_array.append(list(map(lambda cell: float(cell.value), row)))

    vals1 = np.array(vals_array)

    lumin = vals1[0:122,1]

    #luminosity

    luminosity = 0

    for i in range(0, len(lam)):
        luminosity = luminosity + 683*10**(-3)*lumin[i]*total[i]

    # print(luminosity)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam, total

def halley_calc(orbitalpoint):
    aph = 35.08
    per = 0.586
    ecc = 0.967
    a = 0.5*(aph+per)

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)

    #redefining oribtal parameters for orbital plot - otherwise plot is too extreme
    aph = 35.08
    per = 30
    ecc = 0.73
    a = 0.5 * (aph + per)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, r2, ext_ter_spec, lam

def europa_calc(orbitalpoint):
    a = 5.2026
    ecc = 0.048498

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam

def ceres_calc(orbitalpoint):
    a = 2.7675
    ecc = 0.075823

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam

def pluto_calc(orbitalpoint):
    a = 39.48
    ecc = 0.2488

    sol, lam_bb, maximum, r, solar_const, ext_ter_spec, lam = black_body(a, ecc, orbitalpoint)
    r_plot, theta, r2 = orbital_plot_calc(a, ecc, orbitalpoint)

    return sol, lam_bb, maximum, r, solar_const, r_plot, theta, ext_ter_spec, lam
